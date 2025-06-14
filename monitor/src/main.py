import re
import os
import pymupdf
from openpyxl import load_workbook
from docx import Document
from PIL import Image
import pytesseract
import io
import zipfile
from sentence_transformers import SentenceTransformer, util
#import spacy
from pymongo import MongoClient, ReturnDocument
from bson.objectid import ObjectId
import numpy as np
import grpc
from concurrent import futures
import monitor_pb2
import monitor_pb2_grpc
from grpc_health.v1 import health, health_pb2, health_pb2_grpc
import faiss


INDEX_PATH = '/var/faiss/faiss_index.index'

background_executor = futures.ThreadPoolExecutor(max_workers=15)

db_client = MongoClient(os.getenv("MONGO_URI"))
events_collection = db_client["proxyGPT"]["events"]
regex_collection = db_client["proxyGPT"]["regex_rules"]
topics_collection = db_client["proxyGPT"]["topic_rules"]
counter_collection = db_client["proxyGPT"]["faiss_id_counters"]

#nlp = spacy.load("en_core_web_sm")

embeddings_model = SentenceTransformer('all-MiniLM-L6-v2')

def get_next_faiss_id():
    counter = counter_collection.find_one_and_update(
        {"_id": "faiss_id_counter"},
        {"$inc": {"last_id": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER
    )
    return counter["last_id"]


def chunk_text(text, chunk_size=500, overlap=50):
    """Split text into chunks of `chunk_size` with optional `overlap`."""
    words = text.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        chunk = ' '.join(words[i:i + chunk_size])
        chunks.append(chunk)
    return chunks



def analyze_topic_leak(text):

    leaked_topics = []
    chunk_embeddings = obtain_embeddings_from_text(text)
    
    # Let's check if the is similarity with the faiss index
    embedding_vectors = np.array([emb['embedding'] for emb in chunk_embeddings], dtype='float32')
    for embedding in embedding_vectors:
        #faiss_index.hnsw.efSearch = 16  # Query time accuracy/speed tradeoff, default is 16
        scores, indices = faiss_index.search(embedding.reshape(1, -1), 10)  # Search for top 10 similar embeddings
        for score, idx in zip(scores[0], indices[0]):
            if score >= 0.3:        #Cosine similarity threshold higher or equals to 0.3
                print(f"Found similar embedding with score {score} at index {idx}")

                doc = topics_collection.find_one(
                    {"faiss_indexes.faiss_id": int(idx)},
                    {"faiss_indexes.$": 1, "name": 1}  # Only project the matched index with the chunk
                )

                
                if doc:
                    print("Matched Document ID:", doc["_id"])
                    leaked_topics.append({"name": doc['name'], "faiss_id": int(idx), "score": float(score)})

                else:
                    print("No matching document found.")


    return leaked_topics

    
def obtain_embeddings_from_text(text):
    chunks = chunk_text(text)
    embeddings = embeddings_model.encode(chunks, normalize_embeddings=True)

    linked_data = [{"chunk": chunk, "embedding": embedding.tolist()} for chunk, embedding in zip(chunks, embeddings)]

    return linked_data



def analyze_text(text):
    leak = {}
    leak['regex'] = analyze_text_regex(text)
    leak['topic'] = analyze_topic_leak(text)

    return leak

def analyze_text_regex(text):

    regexes = {}
    for doc in regex_collection.find():
        for regex_name, regex_value in doc.items():
            if regex_name != "_id":
                compiled_regex = re.compile(regex_value)
                for line in text.splitlines():
                    match = re.search(compiled_regex, line)
                    if match:
                        regexes[match.group()] = regex_name
        
    return regexes

def decode_file(filepath, content_type):

    text = ""
    if content_type == "application/pdf":
        with pymupdf.open(filepath) as doc:  # open document
            text = chr(12).join([page.get_text() for page in doc])

            #Extract images and apply OCR
            for page_num in range(len(doc)):
                page = doc[page_num]
                image_list = page.get_images(full=True)
                print(f"[+] Found {len(image_list)} images on page {page_num}")

                for img_index, img in enumerate(image_list):
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    image_bytes = base_image["image"]
                    image = Image.open(io.BytesIO(image_bytes))
                    text += pytesseract.image_to_string(image)
                                    
    elif content_type == "application/vnd.ms-excel" or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        workbook = load_workbook(filename=filepath)
        text_data = []
    
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = ' '.join([str(cell) for cell in row if cell is not None])
                text_data.append(row_text)
        
        text = '\n'.join(text_data)

    elif content_type == "application/msword" or content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        doc = Document(filepath)
        text = [para.text for para in doc.paragraphs if para.text.strip()]
        text = '\n'.join(text)

        extract_images_from_docx(filepath)


    elif content_type == "image/jpeg" or content_type == "image/png" or content_type == "image/gif" or content_type == "image/bmp" or content_type == "image/webp" or content_type == "image/svg+xml" or content_type == "image/tiff" or content_type == "image/vnd.microsoft.icon":
        
        image = Image.open(filepath)
        text = pytesseract.image_to_string(image)

    elif content_type == "text/plain":
        with open(filepath, 'r') as file:
            text = file.read()
    
    return text


def analyze_file(filepath, content_type):
    text = decode_file(filepath, content_type)
    return text, analyze_text(text)



def extract_images_from_docx(docx_path, output_folder="extracted_images"):
    with zipfile.ZipFile(docx_path, 'r') as docx_zip:
        # Create output folder if it doesn't exist
        os.makedirs(output_folder, exist_ok=True)

        # Loop through files in the ZIP and extract images
        for file in docx_zip.namelist():
            if file.startswith("word/media/"):
                filename = os.path.basename(file)
                if filename:  # skip folders
                    target_path = os.path.join(output_folder, filename)
                    with open(target_path, "wb") as img_file:
                        img_file.write(docx_zip.read(file))
                    print(f"Saved image: {target_path}")


def on_event_added(event_id):
    try:
        print(f"Started on_event_added for Event ID: {event_id}")
        event = events_collection.find_one({'_id': ObjectId(event_id)})

        print(f"Event obtained")
    
        leak = {}
        result = {}

        if event['rational'] == "Conversation":
            print(f"Conversation, analysing: {event['content']}")
            leak = analyze_text(event['content'])
            print(f"Done: {leak}")


            result = events_collection.update_one(
                {"_id": ObjectId(event_id)},
                {"$set": {"leak": leak}}
            )

        elif event['rational'] == "Attached file":
            text, leak = analyze_file(event['filepath'], event['content_type'])

            result = events_collection.update_one(
                {"_id": ObjectId(event_id)},
                {"$set": {"leak": leak}}
            )


        if result.modified_count > 0:
            print("Document updated successfully.")
        else:
            print("No changes made or document not found.")


        print(f"Finished long task for Event ID: {event_id}")

    except Exception as e:
        # Handle the exception
        print(f"An error occurred: {e}")


def on_topic_rule_added(topic_rule_id):
    try:

        print(f"Started on_topic_rule_added for Topic Rule ID: {topic_rule_id}")
        topic_rule = topics_collection.find_one({'_id': ObjectId(topic_rule_id)})

        print(f"Topic Rule obtained: {topic_rule}")

        chunks = chunk_text(topic_rule['pattern'])
        embeddings = embeddings_model.encode(chunks, normalize_embeddings=True)
        faiss_indexes = [{"chunk": chunk, "faiss_id": get_next_faiss_id()} for chunk in chunks]

        # Prepare data
        ids = np.array([index['faiss_id'] for index in faiss_indexes], dtype='int64')

        #Add embeddings to FAISS index
        faiss_index.add_with_ids(embeddings, ids)

        #Save the FAISS index to disk
        faiss.write_index(faiss_index, INDEX_PATH)

        topics_collection.update_one(
            {'_id': ObjectId(topic_rule_id)},  # Filter to find the document
            {'$set': {'faiss_indexes': faiss_indexes}}  # Field to add or update
        )

        #print(f"Topic Rule encoded: {encoded}")

    except Exception as e:
        # Handle the exception
        print(f"An error occurred: {e}")




class MonitorServicer(monitor_pb2_grpc.MonitorServicer):

    def EventAdded(self, request, context):
        print(f"Received Event ID: {request.id}")
        background_executor.submit(on_event_added, request.id)
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)
    
    def TopicRuleAdded(self, request, context):
        print(f"Received Topic Rule ID: {request.id}")
        background_executor.submit(on_topic_rule_added, request.id)
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)
    


def main():

    #Load FAISS index if it exists
    if os.path.exists(INDEX_PATH):
        print("📦 Loading existing FAISS index...")
        global faiss_index
        faiss_index = faiss.read_index(INDEX_PATH)
    else:   #Inizialize it
        print("🔍 No existing FAISS index found, starting fresh...")
        dim = 384   # Dimension of the embeddings (for 'all-MiniLM-L6-v2')
        flat = faiss.IndexFlatIP(dim)   # Use Inner Product (IP) for similarity search
        #flat = faiss.IndexHNSWFlat(dim, 32, faiss.METRIC_INNER_PRODUCT)  # Using Inner Product (cosine similarity with normalized vectors)
        #flat.hnsw.efConstruction = 40  # Construction time/search accuracy tradeoff, default is 40
        faiss_index = faiss.IndexIDMap2(flat) # Create an index with ID mapping

    server = grpc.server(futures.ThreadPoolExecutor(max_workers=10))
    monitor_pb2_grpc.add_MonitorServicer_to_server(MonitorServicer(), server)

    #For health check to ensure proper start up of the containers
    # Add health service
    health_servicer = health.HealthServicer()
    health_pb2_grpc.add_HealthServicer_to_server(health_servicer, server)
    health_servicer.set('', health_pb2.HealthCheckResponse.SERVING)

    server.add_insecure_port("[::]:50051")
    server.start()
    print("Server running on port 50051...")
    server.wait_for_termination() 

if __name__ == "__main__":
    main()